import cv2
import numpy as np
import os
import cfg

from openpyxl import load_workbook
from openpyxl import Workbook

# im = cv2.imdecode(
#     np.fromfile(r"C:\Users\柳博\Downloads\space-tech-remote-sensing-ob-detection-dataset\val\images\P5789.png",
#                 dtype=np.uint8), cv2.IMREAD_COLOR)
#
# print(im.shape)

# print(os.getcwd())
#
# execlname = "play.xlsx"
# if os.path.exists(os.path.join(os.getcwd(), execlname)):
#    wb = load_workbook(execlname)
#    if "total" in wb.sheetnames:
#        ws = wb.get_sheet_by_name("total")
#    else:
#        ws = wb.create_sheet("total")
# else:
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "total"
#
# ws["A1"] = "hello"
# wb.save(execlname)

# wb = Workbook()
# ws = wb.active
#
# filenames = os.listdir(cfg.test_image)
# rownumber = 1
# for filename in filenames:
#     file_id = filename.split(".")[0]
#     ws["A" + str(rownumber)] = file_id
#     rownumber += 1
#
# wb.save("test_id.xlsx")

head = {"file_id": "A",
        "image_width": "B",
        "image_height": "C",
        "image_source": "D",
        "gsd": "E",
        "x1": "F",
        "y1": "G",
        "x2": "H",
        "y2": "I",
        "x3": "J",
        "y3": "K",
        "x4": "L",
        "y4": "M",
        "class_type": "N",
        "difficult": "O"}

wb = Workbook()
ws = wb.active
head_list = [i for i in head.keys()]
for i in range(len(head_list)):
    ws[chr(ord("A")+i) + "1"] = head_list[i]
wb.save("test.xlsx")
