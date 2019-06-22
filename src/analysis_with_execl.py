import seaborn as sns
import matplotlib.pyplot as plt
import os
import pandas as pd
import json
import cfg
from PIL import Image
import cv2
import numpy as np
from tqdm import tqdm

from openpyxl import load_workbook, Workbook


def write_lines_to_execl(worksheet, head, row_number, file_id, images_dir, labelTxt_dir=None):
    image_file = os.path.join(images_dir, file_id + ".png")
    # with Image.open(image_file) as im:
    #     image_width, image_height = im.size
    im = cv2.imdecode(
        np.fromfile(image_file,
                    dtype=np.uint8), cv2.IMREAD_COLOR)
    image_height, image_width = im.shape[0], im.shape[1]
    if labelTxt_dir:
        label_file = os.path.join(labelTxt_dir, file_id + ".txt")
        image_source = ""
        gsd = 0.0
        x1_list = []
        y1_list = []
        x2_list = []
        y2_list = []
        x3_list = []
        y3_list = []
        x4_list = []
        y4_list = []
        class_type_list = []
        difficult_list = []
        with open(label_file) as f:
            lines = f.readlines()
            image_source = lines[0].split(":")[1]
            line_num = 1
            while len(lines[line_num].strip()) == 0:
                line_num += 1
            gsd = float(lines[line_num].strip().split(":")[1]) if \
                lines[line_num].strip().split(":")[1] != "null" and \
                lines[line_num].strip().split(":")[1] != "None" else 0
            for line in lines[line_num + 1:]:
                line = line.strip()
                if len(line) == 0:
                    continue
                x1, x2, y1, y2, x3, x4, y3, y4, class_type, difficult = line.split(" ")
                x1_list.append(float(x1))
                y1_list.append(float(y1))
                x2_list.append(float(x2))
                y2_list.append(float(y2))
                x3_list.append(float(x3))
                y3_list.append(float(y3))
                x4_list.append(float(x4))
                y4_list.append(float(y4))
                class_type_list.append(class_type)
                difficult_list.append(int(difficult))
        for i in range(len(x1_list)):
            worksheet[head["file_id"] + str(row_number)] = file_id
            worksheet[head["image_width"] + str(row_number)] = image_width
            worksheet[head["image_height"] + str(row_number)] = image_height
            worksheet[head["image_source"] + str(row_number)] = image_source
            worksheet[head["gsd"] + str(row_number)] = gsd
            worksheet[head["x1"] + str(row_number)] = x1_list[i]
            worksheet[head["y1"] + str(row_number)] = y1_list[i]
            worksheet[head["x2"] + str(row_number)] = x2_list[i]
            worksheet[head["y2"] + str(row_number)] = y2_list[i]
            worksheet[head["x3"] + str(row_number)] = x3_list[i]
            worksheet[head["y3"] + str(row_number)] = y3_list[i]
            worksheet[head["x4"] + str(row_number)] = x4_list[i]
            worksheet[head["y4"] + str(row_number)] = y4_list[i]
            worksheet[head["class_type"] + str(row_number)] = class_type_list[i]
            worksheet[head["difficult"] + str(row_number)] = difficult_list[i]
            row_number += 1
        return row_number
    else:  # 如果没有标签
        worksheet[head["file_id"] + str(row_number)] = file_id
        worksheet[head["image_width"] + str(row_number)] = image_width
        worksheet[head["image_height"] + str(row_number)] = image_height
        row_number += 1
        return row_number


if __name__ == "__main__":
    # ignore_file_id_list = ["P11054", "P6637", "P3536", "P5203", "P9847"  # 因为图片太大而无法加载 train
    #     , "P5789",  # 验证集
    #                        "P8137", "P6905"]  # 测试集暂时没有
    ignore_file_id_list = []
    head = {"file_id": "A",
            "image_width": "B",
            "image_height": "C",
            "image_source": "D",
            "gsd": "E",
            "x1": "F",
            "y1": "G",
            "x2": "H",
            "y2": "I",
            "x3": "J",
            "y3": "K",
            "x4": "L",
            "y4": "M",
            "class_type": "N",
            "difficult": "O"}

    # 创建工作表
    execl_filename = "dataset.xlsx"
    if os.path.exists(os.path.join(os.getcwd(), execl_filename)):
        wb = load_workbook(execl_filename)
        if "total" in wb.sheetnames:
            worksheet_0 = wb.get_sheet_by_name("total")
        else:
            worksheet_0 = wb.create_sheet(title="total")
        if "train" in wb.sheetnames:
            worksheet_1 = wb.get_sheet_by_name("train")
        else:
            worksheet_1 = wb.create_sheet(title="train")
        if "val" in wb.sheetnames:
            worksheet_2 = wb.get_sheet_by_name("val")
        else:
            worksheet_2 = wb.create_sheet(title="val")
        if "test" in wb.sheetnames:
            worksheet_3 = wb.get_sheet_by_name("test")
        else:
            worksheet_3 = wb.create_sheet(title="test")
    else:
        wb = Workbook()
        worksheet_0 = wb.active
        worksheet_0.title = "total"
        worksheet_1 = wb.create_sheet(title="train")
        worksheet_2 = wb.create_sheet(title="val")
        worksheet_3 = wb.create_sheet(title="test")

    print(wb.sheetnames)

    # 文件数量
    train_image = os.listdir(cfg.train_image)
    train_labelTxt = os.listdir(cfg.train_labelTxt)
    val_image = os.listdir(cfg.val_image)
    val_labelTxt = os.listdir(cfg.val_labelTxt)
    test_image = os.listdir(cfg.test_image)
    # total
    row_number = 1
    image_sets = [train_image, val_image, test_image]
    images_dirs = [cfg.train_image, cfg.val_image, cfg.test_image]
    for i, _ in zip(range(len(image_sets)), tqdm(range(len(image_sets)), desc="创建total表格")):
        for image_file_name in image_sets[i]:
            file_id = image_file_name.split(".")[0]
            if file_id in ignore_file_id_list:
                continue
            row_number = write_lines_to_execl(worksheet_0, head, row_number, file_id, images_dir=images_dirs[i])
            if row_number > 6:
                break
    # train
    row_number = 1
    for image_file_name, _ in zip(train_image, tqdm(range(len(train_image)), desc="创建train表格")):
        file_id = image_file_name.split(".")[0]
        if file_id in ignore_file_id_list:
            continue
        row_number = write_lines_to_execl(worksheet_1, head, row_number, file_id,
                                          images_dir=cfg.train_image, labelTxt_dir=cfg.train_labelTxt)
    # val
    row_number = 1
    for image_file_name, _ in zip(val_image, tqdm(range(len(val_image)), desc="创建val表格")):
        file_id = image_file_name.split(".")[0]
        if file_id in ignore_file_id_list:
            continue
        row_number = write_lines_to_execl(worksheet_2, head, row_number, file_id,
                                          images_dir=cfg.val_image, labelTxt_dir=cfg.val_labelTxt)
    # test
    row_number = 1
    for image_file_name, _ in zip(test_image, tqdm(range(len(test_image)), desc="创建test表格")):
        file_id = image_file_name.split(".")[0]
        if file_id in ignore_file_id_list:
            continue
        row_number = write_lines_to_execl(worksheet_2, head, row_number, file_id,
                                          images_dir=cfg.test_image)

    # 保存
    wb.save(execl_filename)
