import os
from openpyxl import load_workbook

import cfg

execl_filename = os.path.join(os.getcwd(), "..", "execl", "GoogleEarthSmallDataset.xlsx")
wb = load_workbook(execl_filename)
ws = wb.get_sheet_by_name("small_dataset")
print(ws.max_row)
print(ws.max_column)

def copy_image_and_labelTxt():
    for row in ws.rows:
        file_id = row[0].value
        is_train = row[1].value == "train"
        origin_img = os.path.join(cfg.train_image if is_train else cfg.val_image, file_id + ".png")
        origin_labelTxt = os.path.join(cfg.train_labelTxt if is_train else cfg.val_labelTxt, file_id + ".txt")
        small_dataset_img = os.path.join(cfg.small_train_image if is_train else cfg.small_val_image, file_id + ".png")
        small_dataset_labelTxt = os.path.join(cfg.small_train_labelTxt if is_train else cfg.small_val_labelTxt,
                                              file_id + ".txt")
        os.system('copy %s %s' % (origin_img, small_dataset_img))  # 拷文件
        os.system('copy %s %s' % (origin_labelTxt, small_dataset_labelTxt))  # 拷文件
        if not os.path.isfile(small_dataset_img):
            print("copy image field:%s" % file_id)
        if not os.path.isfile(small_dataset_labelTxt):
            print("copy labelTxt field:%s" % file_id)


if __name__ == "__main__":
    copy_image_and_labelTxt()
